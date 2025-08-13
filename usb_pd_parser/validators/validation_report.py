#!/usr/bin/env python3
"""
Validation Report Generator for USB PD Specification Parser

This module provides comprehensive validation of parsing results and generates
detailed Excel reports comparing ToC entries with parsed document sections.
"""

import logging
from typing import List, Dict, Any, Tuple, Optional
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from collections import defaultdict

from ..parsers.toc_parser import TOCEntry
from ..parsers.document_parser import DocumentSection

logger = logging.getLogger(__name__)

class ValidationReport:
    """
    Comprehensive validation system for USB PD specification parsing results.
    
    Compares Table of Contents entries with parsed document sections and generates
    detailed Excel reports with statistics, mismatches, and quality metrics.
    """
    
    def __init__(self, doc_title: str = "USB PD Specification"):
        """
        Initialize the validation report generator.
        
        Args:
            doc_title: Document title for report metadata
        """
        self.doc_title = doc_title
        self.validation_timestamp = datetime.now().isoformat()
        
        # Validation results storage
        self.validation_results = {
            "summary": {},
            "section_comparison": [],
            "missing_sections": [],
            "extra_sections": [],
            "page_mismatches": [],
            "quality_issues": [],
            "statistics": {}
        }
    
    def validate_parsing_results(self, toc_entries: List[TOCEntry], 
                                document_sections: List[DocumentSection]) -> Dict[str, Any]:
        """
        Perform comprehensive validation of parsing results.
        
        Args:
            toc_entries: List of ToC entries from parsing
            document_sections: List of document sections from parsing
            
        Returns:
            Comprehensive validation results dictionary
        """
        logger.info(f"Starting validation of {len(toc_entries)} ToC entries vs {len(document_sections)} document sections")
        
        # Create lookup dictionaries
        toc_map = {entry.section_id: entry for entry in toc_entries}
        section_map = {section.section_id: section for section in document_sections}
        
        # Perform validation checks
        self._validate_section_coverage(toc_map, section_map)
        self._validate_page_consistency(toc_map, section_map)
        self._validate_hierarchical_structure(toc_entries, document_sections)
        self._validate_content_quality(document_sections)
        self._calculate_validation_statistics(toc_entries, document_sections)
        
        # Generate summary
        self._generate_validation_summary()
        
        logger.info("Validation completed successfully")
        return self.validation_results
    
    def _validate_section_coverage(self, toc_map: Dict[str, TOCEntry], 
                                  section_map: Dict[str, DocumentSection]):
        """
        Validate that all ToC sections have corresponding document sections and vice versa.
        
        Args:
            toc_map: Dictionary mapping section IDs to ToC entries
            section_map: Dictionary mapping section IDs to document sections
        """
        # Find missing sections (in ToC but not in document)
        missing_sections = []
        for section_id, toc_entry in toc_map.items():
            if section_id not in section_map:
                missing_sections.append({
                    "section_id": section_id,
                    "title": toc_entry.title,
                    "page": toc_entry.page,
                    "level": toc_entry.level,
                    "issue": "Missing from document sections"
                })
        
        # Find extra sections (in document but not in ToC)
        extra_sections = []
        for section_id, doc_section in section_map.items():
            if section_id not in toc_map:
                extra_sections.append({
                    "section_id": section_id,
                    "title": doc_section.title,
                    "page_start": doc_section.page_start,
                    "level": doc_section.level,
                    "issue": "Not found in ToC"
                })
        
        self.validation_results["missing_sections"] = missing_sections
        self.validation_results["extra_sections"] = extra_sections
        
        logger.info(f"Coverage validation: {len(missing_sections)} missing, {len(extra_sections)} extra sections")
    
    def _validate_page_consistency(self, toc_map: Dict[str, TOCEntry], 
                                  section_map: Dict[str, DocumentSection]):
        """
        Validate page number consistency between ToC and document sections.
        
        Args:
            toc_map: Dictionary mapping section IDs to ToC entries
            section_map: Dictionary mapping section IDs to document sections
        """
        page_mismatches = []
        
        for section_id in set(toc_map.keys()) & set(section_map.keys()):
            toc_entry = toc_map[section_id]
            doc_section = section_map[section_id]
            
            if toc_entry.page != doc_section.page_start:
                page_mismatches.append({
                    "section_id": section_id,
                    "title": toc_entry.title,
                    "toc_page": toc_entry.page,
                    "document_page": doc_section.page_start,
                    "difference": abs(toc_entry.page - doc_section.page_start),
                    "issue": "Page number mismatch"
                })
        
        self.validation_results["page_mismatches"] = page_mismatches
        logger.info(f"Page validation: {len(page_mismatches)} page mismatches found")
    
    def _validate_hierarchical_structure(self, toc_entries: List[TOCEntry], 
                                       document_sections: List[DocumentSection]):
        """
        Validate the hierarchical structure consistency.
        
        Args:
            toc_entries: List of ToC entries
            document_sections: List of document sections
        """
        section_comparison = []
        
        # Create maps for easy lookup
        toc_map = {entry.section_id: entry for entry in toc_entries}
        section_map = {section.section_id: section for section in document_sections}
        
        # Compare entries that exist in both
        for section_id in set(toc_map.keys()) & set(section_map.keys()):
            toc_entry = toc_map[section_id]
            doc_section = section_map[section_id]
            
            # Check for structural inconsistencies
            issues = []
            if toc_entry.level != doc_section.level:
                issues.append(f"Level mismatch: ToC={toc_entry.level}, Doc={doc_section.level}")
            
            if toc_entry.parent_id != doc_section.parent_id:
                issues.append(f"Parent mismatch: ToC={toc_entry.parent_id}, Doc={doc_section.parent_id}")
            
            if toc_entry.title.strip() != doc_section.title.strip():
                issues.append("Title mismatch")
            
            section_comparison.append({
                "section_id": section_id,
                "toc_title": toc_entry.title,
                "doc_title": doc_section.title,
                "toc_page": toc_entry.page,
                "doc_page_start": doc_section.page_start,
                "doc_page_end": doc_section.page_end,
                "toc_level": toc_entry.level,
                "doc_level": doc_section.level,
                "toc_parent": toc_entry.parent_id,
                "doc_parent": doc_section.parent_id,
                "toc_confidence": toc_entry.confidence_score,
                "doc_confidence": doc_section.confidence_score,
                "word_count": doc_section.word_count,
                "content_type": doc_section.content_type,
                "has_tables": doc_section.has_tables,
                "has_figures": doc_section.has_figures,
                "issues": "; ".join(issues) if issues else "No issues",
                "status": "Issues found" if issues else "OK"
            })
        
        self.validation_results["section_comparison"] = section_comparison
        logger.info(f"Structure validation: {len(section_comparison)} sections compared")
    
    def _validate_content_quality(self, document_sections: List[DocumentSection]):
        """
        Validate content quality and identify potential issues.
        
        Args:
            document_sections: List of document sections
        """
        quality_issues = []
        
        for section in document_sections:
            issues = []
            
            # Check for very low confidence scores
            if section.confidence_score < 0.5:
                issues.append(f"Low confidence score: {section.confidence_score:.2f}")
            
            # Check for very short content
            if section.word_count < 10:
                issues.append(f"Very short content: {section.word_count} words")
            
            # Check for empty content
            if not section.content.strip():
                issues.append("Empty content")
            
            # Check for extraction notes (warnings)
            if section.extraction_notes:
                issues.append(f"Extraction warnings: {'; '.join(section.extraction_notes)}")
            
            # Check for unreasonable page ranges
            if section.page_end and section.page_end < section.page_start:
                issues.append("Invalid page range")
            
            if issues:
                quality_issues.append({
                    "section_id": section.section_id,
                    "title": section.title,
                    "page_start": section.page_start,
                    "confidence_score": section.confidence_score,
                    "word_count": section.word_count,
                    "content_length": len(section.content),
                    "issues": "; ".join(issues),
                    "severity": "High" if section.confidence_score < 0.3 or section.word_count < 5 else "Medium"
                })
        
        self.validation_results["quality_issues"] = quality_issues
        logger.info(f"Quality validation: {len(quality_issues)} quality issues found")
    
    def _calculate_validation_statistics(self, toc_entries: List[TOCEntry], 
                                       document_sections: List[DocumentSection]):
        """
        Calculate comprehensive validation statistics.
        
        Args:
            toc_entries: List of ToC entries
            document_sections: List of document sections
        """
        # Basic counts
        toc_count = len(toc_entries)
        doc_count = len(document_sections)
        matched_count = len(set(e.section_id for e in toc_entries) & 
                          set(s.section_id for s in document_sections))
        
        # Level distributions
        toc_levels = defaultdict(int)
        doc_levels = defaultdict(int)
        for entry in toc_entries:
            toc_levels[entry.level] += 1
        for section in document_sections:
            doc_levels[section.level] += 1
        
        # Content type distribution
        content_types = defaultdict(int)
        for section in document_sections:
            content_types[section.content_type] += 1
        
        # Quality metrics
        if document_sections:
            avg_confidence = sum(s.confidence_score for s in document_sections) / len(document_sections)
            avg_word_count = sum(s.word_count for s in document_sections) / len(document_sections)
            total_word_count = sum(s.word_count for s in document_sections)
            total_tables = sum(s.table_count for s in document_sections)
            total_figures = sum(s.figure_count for s in document_sections)
        else:
            avg_confidence = 0.0
            avg_word_count = 0.0
            total_word_count = 0
            total_tables = 0
            total_figures = 0
        
        # Match rates
        toc_match_rate = matched_count / max(1, toc_count)
        doc_match_rate = matched_count / max(1, doc_count)
        overall_match_rate = (toc_match_rate + doc_match_rate) / 2
        
        self.validation_results["statistics"] = {
            "toc_sections_count": toc_count,
            "document_sections_count": doc_count,
            "matched_sections_count": matched_count,
            "missing_sections_count": len(self.validation_results["missing_sections"]),
            "extra_sections_count": len(self.validation_results["extra_sections"]),
            "page_mismatches_count": len(self.validation_results["page_mismatches"]),
            "quality_issues_count": len(self.validation_results["quality_issues"]),
            "toc_match_rate": toc_match_rate,
            "document_match_rate": doc_match_rate,
            "overall_match_rate": overall_match_rate,
            "toc_level_distribution": dict(toc_levels),
            "document_level_distribution": dict(doc_levels),
            "content_type_distribution": dict(content_types),
            "average_confidence_score": avg_confidence,
            "average_word_count": avg_word_count,
            "total_word_count": total_word_count,
            "total_tables": total_tables,
            "total_figures": total_figures
        }
        
        logger.info(f"Statistics calculated: {overall_match_rate:.2%} overall match rate")
    
    def _generate_validation_summary(self):
        """Generate a high-level validation summary."""
        stats = self.validation_results["statistics"]
        
        # Determine overall status
        if stats["overall_match_rate"] >= 0.95 and stats["quality_issues_count"] == 0:
            status = "Excellent"
            color = "green"
        elif stats["overall_match_rate"] >= 0.85 and stats["quality_issues_count"] <= 2:
            status = "Good"
            color = "yellow"
        elif stats["overall_match_rate"] >= 0.70:
            status = "Fair"
            color = "orange"
        else:
            status = "Poor"
            color = "red"
        
        # Generate summary text
        summary_text = f"""
        Validation Summary for {self.doc_title}
        
        Overall Status: {status}
        Match Rate: {stats['overall_match_rate']:.1%}
        
        Section Counts:
        - ToC Sections: {stats['toc_sections_count']}
        - Document Sections: {stats['document_sections_count']}
        - Matched: {stats['matched_sections_count']}
        - Missing: {stats['missing_sections_count']}
        - Extra: {stats['extra_sections_count']}
        
        Quality Metrics:
        - Page Mismatches: {stats['page_mismatches_count']}
        - Quality Issues: {stats['quality_issues_count']}
        - Average Confidence: {stats['average_confidence_score']:.2f}
        - Total Words: {stats['total_word_count']:,}
        """
        
        self.validation_results["summary"] = {
            "status": status,
            "color": color,
            "text": summary_text.strip(),
            "overall_match_rate": stats["overall_match_rate"],
            "validation_timestamp": self.validation_timestamp
        }
    
    def generate_excel_report(self, output_path: str) -> Dict[str, Any]:
        """
        Generate a comprehensive Excel validation report.
        
        Args:
            output_path: Path to save the Excel report
            
        Returns:
            Report generation statistics
        """
        logger.info(f"Generating Excel validation report: {output_path}")
        
        try:
            # Create Excel workbook
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Summary sheet
                self._write_summary_sheet(writer)
                
                # Section comparison sheet
                self._write_section_comparison_sheet(writer)
                
                # Issues sheets
                self._write_issues_sheets(writer)
                
                # Statistics sheet
                self._write_statistics_sheet(writer)
            
            # Apply formatting
            self._apply_excel_formatting(output_path)
            
            report_stats = {
                "report_generated": True,
                "output_path": output_path,
                "sheets_created": ["Summary", "Section_Comparison", "Missing_Sections", 
                                 "Extra_Sections", "Page_Mismatches", "Quality_Issues", "Statistics"],
                "total_rows": sum(len(data) for data in [
                    self.validation_results["section_comparison"],
                    self.validation_results["missing_sections"],
                    self.validation_results["extra_sections"],
                    self.validation_results["page_mismatches"],
                    self.validation_results["quality_issues"]
                ])
            }
            
            logger.info("Excel report generated successfully")
            return report_stats
            
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            return {"report_generated": False, "error": str(e)}
    
    def _write_summary_sheet(self, writer):
        """Write summary information to Excel sheet."""
        summary_data = [
            ["Validation Report", self.doc_title],
            ["Generated", self.validation_timestamp],
            ["Overall Status", self.validation_results["summary"]["status"]],
            ["Overall Match Rate", f"{self.validation_results['summary']['overall_match_rate']:.1%}"],
            [],
            ["Section Counts", ""],
            ["ToC Sections", self.validation_results["statistics"]["toc_sections_count"]],
            ["Document Sections", self.validation_results["statistics"]["document_sections_count"]],
            ["Matched Sections", self.validation_results["statistics"]["matched_sections_count"]],
            ["Missing Sections", self.validation_results["statistics"]["missing_sections_count"]],
            ["Extra Sections", self.validation_results["statistics"]["extra_sections_count"]],
            [],
            ["Quality Metrics", ""],
            ["Page Mismatches", self.validation_results["statistics"]["page_mismatches_count"]],
            ["Quality Issues", self.validation_results["statistics"]["quality_issues_count"]],
            ["Average Confidence", f"{self.validation_results['statistics']['average_confidence_score']:.3f}"],
            ["Total Word Count", f"{self.validation_results['statistics']['total_word_count']:,}"],
            ["Total Tables", self.validation_results["statistics"]["total_tables"]],
            ["Total Figures", self.validation_results["statistics"]["total_figures"]],
        ]
        
        df_summary = pd.DataFrame(summary_data, columns=["Metric", "Value"])
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
    
    def _write_section_comparison_sheet(self, writer):
        """Write section comparison data to Excel sheet."""
        if self.validation_results["section_comparison"]:
            df_comparison = pd.DataFrame(self.validation_results["section_comparison"])
            df_comparison.to_excel(writer, sheet_name="Section_Comparison", index=False)
    
    def _write_issues_sheets(self, writer):
        """Write various issue sheets to Excel."""
        # Missing sections
        if self.validation_results["missing_sections"]:
            df_missing = pd.DataFrame(self.validation_results["missing_sections"])
            df_missing.to_excel(writer, sheet_name="Missing_Sections", index=False)
        
        # Extra sections
        if self.validation_results["extra_sections"]:
            df_extra = pd.DataFrame(self.validation_results["extra_sections"])
            df_extra.to_excel(writer, sheet_name="Extra_Sections", index=False)
        
        # Page mismatches
        if self.validation_results["page_mismatches"]:
            df_mismatches = pd.DataFrame(self.validation_results["page_mismatches"])
            df_mismatches.to_excel(writer, sheet_name="Page_Mismatches", index=False)
        
        # Quality issues
        if self.validation_results["quality_issues"]:
            df_quality = pd.DataFrame(self.validation_results["quality_issues"])
            df_quality.to_excel(writer, sheet_name="Quality_Issues", index=False)
    
    def _write_statistics_sheet(self, writer):
        """Write detailed statistics to Excel sheet."""
        stats = self.validation_results["statistics"]
        
        # Level distribution comparison
        level_data = []
        all_levels = set(stats["toc_level_distribution"].keys()) | set(stats["document_level_distribution"].keys())
        for level in sorted(all_levels):
            level_data.append({
                "Level": level,
                "ToC_Count": stats["toc_level_distribution"].get(level, 0),
                "Document_Count": stats["document_level_distribution"].get(level, 0)
            })
        
        df_levels = pd.DataFrame(level_data)
        df_levels.to_excel(writer, sheet_name="Statistics", index=False, startrow=0)
        
        # Content type distribution
        content_data = []
        for content_type, count in stats["content_type_distribution"].items():
            content_data.append({"Content_Type": content_type, "Count": count})
        
        df_content = pd.DataFrame(content_data)
        df_content.to_excel(writer, sheet_name="Statistics", index=False, startrow=len(level_data) + 3)
    
    def _apply_excel_formatting(self, file_path: str):
        """Apply formatting to the Excel report."""
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Format each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format headers
                if sheet.max_row > 0:
                    for cell in sheet[1]:  # First row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
            
            # Highlight status in summary sheet
            if "Summary" in workbook.sheetnames:
                summary_sheet = workbook["Summary"]
                status_cell = None
                for row in summary_sheet.iter_rows():
                    if row[0].value == "Overall Status":
                        status_cell = row[1]
                        break
                
                if status_cell:
                    status = status_cell.value
                    if status == "Excellent":
                        status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif status == "Good":
                        status_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    elif status == "Fair":
                        status_cell.fill = PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid")
                    else:  # Poor
                        status_cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
            
            workbook.save(file_path)
            logger.info("Excel formatting applied successfully")
            
        except Exception as e:
            logger.warning(f"Failed to apply Excel formatting: {e}")
    
    def get_validation_summary(self) -> Dict[str, Any]:
        """
        Get a concise validation summary for quick assessment.
        
        Returns:
            Dictionary with key validation metrics
        """
        if not self.validation_results.get("summary"):
            return {"error": "Validation not yet performed"}
        
        return {
            "status": self.validation_results["summary"]["status"],
            "overall_match_rate": self.validation_results["summary"]["overall_match_rate"],
            "total_sections_toc": self.validation_results["statistics"]["toc_sections_count"],
            "total_sections_parsed": self.validation_results["statistics"]["document_sections_count"],
            "matched_sections": self.validation_results["statistics"]["matched_sections_count"],
            "missing_sections": self.validation_results["statistics"]["missing_sections_count"],
            "quality_issues": self.validation_results["statistics"]["quality_issues_count"],
            "validation_timestamp": self.validation_timestamp
        }